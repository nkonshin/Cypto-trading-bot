"""
Excel-экспорт результатов бэктеста.
Создаёт структурированный .xlsx файл с детальной информацией по каждой сделке.
"""

import io
import logging
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

from backtesting.backtest import BacktestResult

# Описания метрик для комментариев в ячейках
METRIC_TOOLTIPS = {
    "Win Rate": "Процент прибыльных сделок от общего числа.\n>50% — хороший показатель для большинства стратегий.",
    "PnL (USDT)": "Profit and Loss — общая прибыль или убыток в USDT.",
    "PnL (%)": "Прибыль/убыток в процентах от начального баланса.",
    "Просадка (%)": "Максимальное падение баланса от пикового значения.\nПоказывает наихудший период. <10% — отлично, <20% — приемлемо.",
    "Профит-фактор": "Сумма прибыльных сделок / сумма убыточных.\n>1 — прибыльно, >1.5 — хорошо, >2 — отлично.",
    "Sharpe Ratio": "Доходность с учётом риска.\n>0.5 — приемлемо, >1 — хорошо, >2 — отлично.\nУчитывает волатильность доходности.",
    "Лучшая сделка": "Максимальная прибыль по одной сделке в USDT.",
    "Худшая сделка": "Максимальный убыток по одной сделке в USDT.",
    "Убытков подряд": "Максимальное количество убыточных сделок подряд.\nПоказывает устойчивость стратегии к серии неудач.",
    "Сделок": "Общее количество завершённых сделок за период.",
    "Направление": "Покупка (лонг) — ставка на рост цены.\nПродажа (шорт) — ставка на падение.",
    "Stop Loss": "Цена автоматического закрытия при убытке.\nЗащищает от больших потерь.",
    "Take Profit": "Цена автоматического закрытия при прибыли.\nФиксирует прибыль.",
    "Плечо": "Кредитное плечо. 5x = позиция в 5 раз больше вложенных средств.\nУвеличивает и прибыль, и убыток.",
    "R:R": "Risk/Reward — соотношение риска к потенциальной прибыли.\n1:2 значит TP в 2 раза дальше SL.",
    "Результат": "Прибыль — сделка закрыта с плюсом.\nУбыток — закрыта с минусом.\nБезубыток — около нуля.",
}


def _add_header_comments(ws, headers, row=1):
    """Добавляет комментарии-подсказки к заголовкам."""
    for col, header in enumerate(headers, 1):
        tooltip = METRIC_TOOLTIPS.get(header)
        if tooltip:
            ws.cell(row=row, column=col).comment = Comment(tooltip, "Crypto Bot")

logger = logging.getLogger(__name__)

# Стили
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
RED_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
SUMMARY_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 30)


def export_single_result(result: BacktestResult) -> bytes:
    """Экспортирует один BacktestResult в xlsx-файл (bytes)."""
    wb = Workbook()

    # --- Лист 1: Сводка ---
    ws_summary = wb.active
    ws_summary.title = "Сводка"
    _write_summary_sheet(ws_summary, result)

    # --- Лист 2: Сделки ---
    ws_trades = wb.create_sheet("Сделки")
    _write_trades_sheet(ws_trades, result)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_comparison(results: list[BacktestResult]) -> bytes:
    """Экспортирует сравнение нескольких стратегий в xlsx-файл (bytes)."""
    wb = Workbook()

    # --- Лист 1: Сравнение ---
    ws_compare = wb.active
    ws_compare.title = "Сравнение"
    _write_comparison_sheet(ws_compare, results)

    # --- Листы по каждой стратегии ---
    for r in results:
        safe_name = r.strategy[:28]  # ограничение Excel на длину имени листа
        ws = wb.create_sheet(safe_name)
        _write_trades_sheet(ws, r)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_summary_sheet(ws, result: BacktestResult):
    """Лист со сводной информацией по стратегии."""
    rows = [
        ("Параметр", "Значение"),
        ("Стратегия", result.strategy),
        ("Пара", result.symbol),
        ("Таймфрейм", result.timeframe),
        ("Период", result.period),
        ("", ""),
        ("Начальный баланс", f"{result.initial_balance:.2f} USDT"),
        ("Конечный баланс", f"{result.final_balance:.2f} USDT"),
        ("PnL", f"{result.total_pnl:+.2f} USDT ({result.total_pnl_pct:+.1f}%)"),
        ("", ""),
        ("Всего сделок", result.total_trades),
        ("Прибыльных", result.winning_trades),
        ("Убыточных", result.losing_trades),
        ("Win Rate", f"{result.win_rate:.1f}%"),
        ("", ""),
        ("Лучшая сделка", f"{result.best_trade:+.2f} USDT"),
        ("Худшая сделка", f"{result.worst_trade:+.2f} USDT"),
        ("Средний PnL/сделка", f"{result.avg_pnl_per_trade:+.2f} USDT"),
        ("", ""),
        ("Макс. просадка", f"{result.max_drawdown_pct:.1f}%"),
        ("Макс. убытков подряд", result.max_consecutive_losses),
        ("Профит-фактор", f"{result.profit_factor:.2f}"),
        ("Sharpe Ratio", f"{result.sharpe_ratio:.2f}"),
    ]

    for row_idx, (param, value) in enumerate(rows, 1):
        ws.cell(row=row_idx, column=1, value=param).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=value).border = THIN_BORDER
        if row_idx == 1:
            ws.cell(row=row_idx, column=1).font = HEADER_FONT
            ws.cell(row=row_idx, column=1).fill = HEADER_FILL
            ws.cell(row=row_idx, column=2).font = HEADER_FONT
            ws.cell(row=row_idx, column=2).fill = HEADER_FILL
        elif param == "":
            continue
        else:
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            ws.cell(row=row_idx, column=1).fill = SUMMARY_FILL

    _auto_width(ws)


def _write_trades_sheet(ws, result: BacktestResult):
    """Лист с детальной информацией по каждой сделке."""
    headers = [
        "№", "Направление", "Вход (время)", "Выход (время)",
        "Длительность", "Цена входа", "Цена выхода",
        "Stop Loss", "Take Profit", "Плечо",
        "Объём", "PnL (USDT)", "PnL (%)",
        "Причина входа", "Причина выхода", "Результат",
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header(ws, 1, len(headers))

    for i, t in enumerate(result.trades, 1):
        row = i + 1
        side_ru = "Покупка" if t.side == "buy" else "Продажа"

        # Длительность
        duration = ""
        if t.entry_time and t.exit_time:
            try:
                fmt = "%Y-%m-%d %H:%M"
                dt_entry = datetime.strptime(t.entry_time, fmt)
                dt_exit = datetime.strptime(t.exit_time, fmt)
                delta = dt_exit - dt_entry
                hours = delta.total_seconds() / 3600
                if hours < 1:
                    duration = f"{int(delta.total_seconds() / 60)} мин"
                elif hours < 24:
                    duration = f"{hours:.1f} ч"
                else:
                    duration = f"{delta.days} д {int(hours % 24)} ч"
            except (ValueError, TypeError):
                duration = ""

        outcome = "Прибыль" if t.pnl > 0 else "Убыток" if t.pnl < 0 else "Безубыток"

        values = [
            i, side_ru, t.entry_time or "", t.exit_time or "",
            duration, t.entry_price, t.exit_price,
            t.stop_loss, t.take_profit, t.leverage,
            round(t.amount, 6), round(t.pnl, 2), round(t.pnl_pct, 2),
            t.reason_entry, t.reason_exit, outcome,
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        # Подсветка строки
        fill = GREEN_FILL if t.pnl > 0 else RED_FILL if t.pnl < 0 else None
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = fill

    # Итоговая строка
    if result.trades:
        total_row = len(result.trades) + 2
        ws.cell(row=total_row, column=1, value="ИТОГО").font = Font(bold=True)
        ws.cell(row=total_row, column=11, value=f"Сделок: {result.total_trades}").font = Font(bold=True)
        ws.cell(row=total_row, column=12, value=round(result.total_pnl, 2)).font = Font(bold=True)
        ws.cell(row=total_row, column=13, value=round(result.total_pnl_pct, 2)).font = Font(bold=True)
        for col in range(1, len(headers) + 1):
            ws.cell(row=total_row, column=col).border = THIN_BORDER
            ws.cell(row=total_row, column=col).fill = SUMMARY_FILL

    _auto_width(ws)


def _write_comparison_sheet(ws, results: list[BacktestResult]):
    """Лист сравнения стратегий."""
    headers = [
        "Место", "Стратегия", "Таймфрейм", "Сделок",
        "Win Rate", "PnL (USDT)", "PnL (%)",
        "Просадка (%)", "Профит-фактор", "Sharpe Ratio",
        "Лучшая сделка", "Худшая сделка", "Убытков подряд",
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header(ws, 1, len(headers))

    sorted_results = sorted(results, key=lambda r: r.total_pnl_pct, reverse=True)

    for i, r in enumerate(sorted_results, 1):
        row = i + 1
        values = [
            i, r.strategy, r.timeframe, r.total_trades,
            f"{r.win_rate:.1f}%", round(r.total_pnl, 2), f"{r.total_pnl_pct:+.1f}%",
            f"{r.max_drawdown_pct:.1f}%", round(r.profit_factor, 2), round(r.sharpe_ratio, 2),
            f"{r.best_trade:+.2f}", f"{r.worst_trade:+.2f}", r.max_consecutive_losses,
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        fill = GREEN_FILL if r.total_pnl > 0 else RED_FILL if r.total_pnl < 0 else None
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = fill

    # Глоссарий метрик под таблицей
    glossary = [
        ("Win Rate", "Процент прибыльных сделок от общего числа. >50% — хороший показатель."),
        ("PnL", "Profit and Loss — общая прибыль или убыток. В USDT и в % от начального баланса."),
        ("Просадка", "Максимальное падение баланса от пика. <10% — отлично, <20% — приемлемо, >30% — рискованно."),
        ("Профит-фактор", "Сумма прибыли / сумма убытков. >1 — прибыльно, >1.5 — хорошо, >2 — отлично."),
        ("Sharpe Ratio", "Доходность с учётом риска (волатильности). >0.5 — приемлемо, >1 — хорошо, >2 — отлично."),
        ("Лучшая / Худшая сделка", "Максимальная прибыль и максимальный убыток по одной сделке в USDT."),
        ("Убытков подряд", "Макс. серия убыточных сделок подряд. Показывает устойчивость к чёрным полосам."),
    ]

    glossary_start = len(sorted_results) + 4  # 2 строки отступа после таблицы

    # Заголовок глоссария
    cell = ws.cell(row=glossary_start, column=1, value="Глоссарий метрик")
    cell.font = Font(bold=True, size=12)

    glossary_start += 1
    term_font = Font(bold=True, size=10, color="2F5496")
    desc_font = Font(size=10, color="444444")

    for idx, (term, desc) in enumerate(glossary):
        row = glossary_start + idx
        cell_term = ws.cell(row=row, column=1, value=term)
        cell_term.font = term_font
        cell_term.alignment = Alignment(horizontal="left")

        cell_desc = ws.cell(row=row, column=2, value=desc)
        cell_desc.font = desc_font
        cell_desc.alignment = Alignment(horizontal="left", wrap_text=True)
        # Объединяем ячейки 2-6 чтобы описание влезло
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)

    _auto_width(ws)
    # Делаем колонку 1 шире для терминов
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 22)
